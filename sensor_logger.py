import os
import re
import time
import datetime
import zipfile
import smtplib
import ssl
import socket
from collections import defaultdict
from zipfile import ZipFile, ZIP_DEFLATED, ZipInfo
from email.message import EmailMessage
from pymodbus.client import ModbusTcpClient
from openpyxl import load_workbook

# === Ayarlar ===
log_dir = "/home/testonaylab/projeler/sensor"
email_sender = "testonayraspberrypi@gmail.com"
email_password = "myyo tnqh idwl qgkb"  # ÖNERİ: os.environ["SMTP_APP_PASS"] kullanın
MAX_ATTACH_MB = 20  # Gmail pratik limit (~25MB) altında tut

gazlab_mail = "fatih.cilesiz@beko.com"
elektriklab_mail = "ferhat_bicer@beko.com"

sensors = {
    "GazLab": {
        "ip": "10.114.8.251",
        "registers": [48, 49, 50],
        "email": gazlab_mail
    },
    "ElektrikLab": {
        "ip": "10.114.8.223",
        "registers": [48, 49, 50],
        "email": elektriklab_mail
    },
    "FırınPerformansLab": {
        "ip": "10.114.8.136",
        "registers": [48, 49, 50],
        "email": elektriklab_mail
    }
}

# Backlog bir kez çalıştırma bayrağı
BACKLOG_PROCESSED = False

# === Yardımcılar ===
def lan_status_multi():
    # 53 ve 443 portlarından birini başarıyla açabilirse internet var say
    for host, port in [("8.8.8.8", 53), ("1.1.1.1", 53), ("www.google.com", 443)]:
        try:
            socket.setdefaulttimeout(3)
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.connect((host, port)); s.close()
            return True
        except Exception:
            continue
    return False

def read_values(ip, registers):
    values = []
    client = ModbusTcpClient(host=ip, port=502)
    try:
        if not client.connect():
            raise ConnectionError(f"{ip} bağlantı başarısız")
        for reg in registers:
            result = client.read_holding_registers(address=reg, count=1, slave=1)
            if not result.isError():
                raw = result.registers[0]
                values.append(round(raw / 10.0, 1))
            else:
                values.append(None)
    except Exception as e:
        print(f"[HATA] {ip} için veri okunamadı → {e}")
        values = [None] * len(registers)
    finally:
        client.close()
    return values

def log_data(lab_name, values):
    today = time.strftime("%d-%m-%Y")
    filename = f"sensor_log_{today}_{lab_name}.txt"
    filepath = os.path.join(log_dir, filename)
    if not os.path.exists(filepath):
        with open(filepath, 'w') as f:
            f.write("Zaman\tSıcaklık (°C)\tNem (%)\tYoğuşma Noktasi (°C)\n")
    now = time.strftime("%H:%M:%S")
    formatted = [str(v) if v is not None else "None" for v in values]
    line = f"{now}\t{formatted[0]}\t{formatted[1]}\t{formatted[2]}\n"
    with open(filepath, 'a') as f:
        f.write(line)
    print(f"{lab_name} → {line.strip()}")

# TXT dosya adı: sensor_log_DD-MM-YYYY_LAB.txt
LOG_TXT_PAT = re.compile(r"^sensor_log_(\d{2}-\d{2}-\d{4})_(.+)\.txt$")

def parse_date(dmy: str) -> datetime.date:
    return datetime.datetime.strptime(dmy, "%d-%m-%Y").date()

def iso_year_week(dt: datetime.date):
    y, w, _ = dt.isocalendar()
    return y, w

# === EXCEL DOLUM ===
def populate_excel_template(txt_path, lab_name, output_path,
                            template_filename="SICAKLIK ÖLÇÜM.xlsx"):
    # TXT oku
    with open(txt_path, 'r') as f:
        lines = f.readlines()
    if len(lines) < 2:
        print(f"[Atlandı] Boş/başlıksız log: {txt_path}")
        return False

    base = os.path.basename(txt_path)
    m = LOG_TXT_PAT.match(base)
    if not m:
        print(f"[Uyarı] Tarih dosya adından çözülemedi: {base}")
        file_date = ""
    else:
        file_date = m.group(1)

    data_rows = []
    for raw in lines[1:]:
        parts = raw.rstrip("\n").split("\t")
        if len(parts) < 4: parts = parts + ["None"]*(4-len(parts))
        data_rows.append(parts[:4])

    template_path = os.path.join(log_dir, template_filename)
    wb = load_workbook(template_path)
    ws = wb.active

    # Başlıklar
    ws["C1"] = lab_name         # LAB_NAME yerine
    ws["D1"] = file_date        # DATE yerine (GG-AA-YYYY)

    # A3↓ zaman, B3↓ sıcaklık, C3↓ %RH, D3↓ çiy noktası
    start = 3
    for i, row in enumerate(data_rows):
        r = start + i
        tm, temp, rh, dew = row
        ws.cell(row=r, column=1, value=tm)
        ws.cell(row=r, column=2, value=(float(temp) if temp not in ("None", "", None) else None))
        ws.cell(row=r, column=3, value=(float(rh)   if rh   not in ("None", "", None) else None))
        ws.cell(row=r, column=4, value=(float(dew)  if dew  not in ("None", "", None) else None))

    wb.save(output_path)
    return True

def txts_for_date(date_obj: datetime.date):
    """Verilen tarih için klasördeki (txt_path, lab_name) çiftlerini döndürür."""
    out = []
    wanted = date_obj.strftime("%d-%m-%Y")
    for fname in os.listdir(log_dir):
        m = LOG_TXT_PAT.match(fname)
        if not m: continue
        dmy, lab = m.groups()
        if dmy == wanted:
            out.append((os.path.join(log_dir, fname), lab))
    return sorted(out)

def ensure_excel_for_txt(txt_path: str, lab_name: str) -> str | None:
    """TXT → XLSX (gerekliyse). Başarılıysa XLSX yolunu döndürür."""
    base, _ = os.path.splitext(os.path.basename(txt_path))
    xlsx_path = os.path.join(log_dir, base + ".xlsx")
    try:
        need = True
        if os.path.exists(xlsx_path):
            need = os.path.getmtime(txt_path) > os.path.getmtime(xlsx_path)
        if need:
            ok = populate_excel_template(txt_path, lab_name, xlsx_path,
                                         template_filename="SICAKLIK ÖLÇÜM.xlsx")
            if not ok:
                print(f"[Atlandı] Boş/başlıksız TXT: {txt_path}")
                return None
        return xlsx_path if os.path.exists(xlsx_path) else None
    except Exception as e:
        print(f"[HATA] Excel oluşturulamadı ({txt_path}): {e}")
        return None

# === HAFTALIK ZIP (yalnızca XLSX) ===
def zip_week_excels(lab: str, year: int, week: int) -> list[str]:
    """
    Verilen LAB, ISO yıl/hafta için klasördeki **sadece Excel** dosyalarını zipler.
    20MB üstünde _partN olarak böler. Dönen liste: oluşturulan zip yolları.
    """
    week_excels = []
    for fname in os.listdir(log_dir):
        if not (fname.endswith(".xlsx") and fname.startswith("sensor_log_")):
            continue
        m = LOG_TXT_PAT.match(fname.replace(".xlsx", ".txt"))
        if not m: 
            continue
        dmy, lab_name = m.groups()
        if lab_name != lab: 
            continue
        dt = parse_date(dmy)
        y, w = iso_year_week(dt)
        if (y, w) == (year, week):
            week_excels.append(fname)

    if not week_excels:
        return []

    week_excels.sort(key=lambda n: parse_date(LOG_TXT_PAT.match(n.replace(".xlsx",".txt")).group(1)))

    created = []
    part = 1
    tmp_zip = os.path.join(log_dir, f"week_{year}_{week}_{lab}.zip")
    zf = ZipFile(tmp_zip, "w", compression=ZIP_DEFLATED)

    try:
        for x in week_excels:
            fpath = os.path.join(log_dir, x)
            zf.write(fpath, arcname=x)
            zf.fp.flush()
            size_mb = os.path.getsize(tmp_zip) / (1024*1024)
            if size_mb > MAX_ATTACH_MB:
                zf.close()
                final = os.path.join(log_dir, f"week_{year}_{week}_{lab}_part{part}.zip")
                os.rename(tmp_zip, final); created.append(final)
                part += 1
                tmp_zip = os.path.join(log_dir, f"week_{year}_{week}_{lab}.zip")
                zf = ZipFile(tmp_zip, "w", compression=ZIP_DEFLATED)
                zf.write(fpath, arcname=x)

        zf.close()
        final_size = os.path.getsize(tmp_zip) / (1024*1024)
        if final_size > 0:
            if part > 1:
                final = os.path.join(log_dir, f"week_{year}_{week}_{lab}_part{part}.zip")
                os.rename(tmp_zip, final); created.append(final)
            else:
                created.append(tmp_zip)
        else:
            os.remove(tmp_zip)
    except Exception as e:
        try: zf.close()
        except: pass
        print(f"[HATA] Ziplenirken: {lab} W{week} {year} → {e}")
    return created

# === MAIL GÖNDERİM (week_*.zip) ===
LAB_ALIASES = {
    "gazlab": "GazLab",
    "elektriklab": "ElektrikLab",
    "firinperformanslab": "FırınPerformansLab",
    "fırınperformanslab": "FırınPerformansLab",
}

def normalize_lab(name: str) -> str | None:
    lowered = (name.replace("İ","i").replace("I","i").replace("ı","i").lower())
    lowered = re.sub(r"[^a-z0-9]+", "", lowered)
    return LAB_ALIASES.get(lowered, None)

def send_pending_zip_files():
    if not lan_status_multi():
        print("[Uyarı] İnternet bağlantısı yok (53/443). ZIP gönderimi ertelendi.")
        return

    zip_pat = re.compile(r"^week_(\d{4})_(\d{1,2})_(.+?)(?:_part\d+)?\.zip$", re.IGNORECASE)

    for file in sorted(os.listdir(log_dir)):
        if not (file.endswith(".zip") and file.startswith("week_")):
            continue

        m = zip_pat.match(file)
        if not m:
            print(f"[Atlandı] Beklenen ada uymuyor: {file}")
            continue

        year, week, lab_raw = m.group(1), m.group(2), m.group(3)
        zip_path = os.path.join(log_dir, file)

        # Boyut kontrolü
        size_mb = os.path.getsize(zip_path) / (1024 * 1024)
        if size_mb > MAX_ATTACH_MB:
            print(f"[Uyarı] {file} {size_mb:.1f} MB > {MAX_ATTACH_MB} MB; parçalara bölünmüş olmalıydı.")

        # Lab eşleştirme
        lab_key = normalize_lab(lab_raw)
        if lab_key is None or lab_key not in sensors:
            recipient = elektriklab_mail
            subject_lab_name = lab_raw
            print(f"[Uyarı] '{lab_raw}' sensors'ta yok. Fallback: {recipient}")
        else:
            recipient = sensors[lab_key]['email']
            subject_lab_name = lab_key

        msg = EmailMessage()
        msg['Subject'] = f"Haftalık Log - {subject_lab_name} - Hafta {week}, {year}"
        msg['From'] = email_sender
        msg['To'] = recipient
        msg.set_content(f"{subject_lab_name} logları ekte .zip formatında yer almaktadır.")

        with open(zip_path, 'rb') as f:
            msg.add_attachment(f.read(), maintype='application', subtype='zip', filename=file)

        try:
            context = ssl.create_default_context()
            with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
                smtp.set_debuglevel(1)  # sunucu diyaloğunu göster
                smtp.login(email_sender, email_password)
                smtp.send_message(msg)
                print(f"[OK] {subject_lab_name} → ZIP mail gönderildi → {file}")

            # Başarılıysa ZIP’i sil
            os.remove(zip_path)
            print(f"[Temizlendi] ZIP silindi: {file}")

        except smtplib.SMTPResponseException as e:
            print(f"[SMTP] code={e.smtp_code} msg={e.smtp_error} dosya={file}")
        except Exception as e:
            print(f"[Genel Hata] {file} gönderilemedi → {e}")

# === GÜN SONU / HAFTA SONU / TEMİZLİK ===
def daily_finalize_to_excel(yesterday: datetime.date):
    """Dünün TXT → Excel dönüşümleri (eksik kalmışsa hepsi)."""
    pairs = txts_for_date(yesterday)
    if not pairs:
        print(f"[GünSonu] {yesterday} için TXT yok.")
        return
    for txt_path, lab in pairs:
        ensure_excel_for_txt(txt_path, lab)
    print(f"[GünSonu] {yesterday} için Excel üretimi tamam.")

def weekly_finalize_and_send_if_sunday00(now_next_minute: datetime.datetime):
    """
    Pazar 00:00 anında:
    - Dünün (Cumartesi) ISO haftası için her LAB'ın Excel’lerini zipler,
    - ZIP’leri gönderir,
    - Gönderim başarılı olan haftanın Excel’lerini temizler.
    """
    if not (now_next_minute.weekday() == 6 and now_next_minute.hour == 0 and now_next_minute.minute == 0):
        return

    # Önce dünün (Cumartesi) gün sonu Excel'leri üretildi (main loop'ta daily_finalize çağırıyoruz)
    yesterday = (now_next_minute.date() - datetime.timedelta(days=1))
    y, w = iso_year_week(yesterday)

    any_zip = False
    for lab in sensors.keys():
        zips = zip_week_excels(lab, y, w)
        if zips:
            any_zip = True
            print(f"[HaftaSonu] {lab} için {y}-W{w} zip hazır: {len(zips)} parça.")

    if any_zip:
        send_pending_zip_files()
        # Başarılı olan ZIP'ler silindi; şimdi haftanın Excel’lerini temizle
        for fname in list(os.listdir(log_dir)):
            if not (fname.endswith(".xlsx") and fname.startswith("sensor_log_")):
                continue
            m = LOG_TXT_PAT.match(fname.replace(".xlsx",".txt"))
            if not m: 
                continue
            dmy, lab = m.groups()
            dt = parse_date(dmy)
            yy, ww = iso_year_week(dt)
            if (yy, ww) == (y, w):
                try:
                    os.remove(os.path.join(log_dir, fname))
                except Exception as e:
                    print(f"[Uyarı] Excel silinemedi ({fname}): {e}")

def retention_cleanup(now_next_minute: datetime.datetime):
    """
    - TXT: son 14 gün dışını sil.
    - ZIP: son 2 ISO hafta dışını sil.
    Şablona dokunma.
    """
    # TXT tutma
    cutoff_txt = now_next_minute.date() - datetime.timedelta(days=14)
    for fname in list(os.listdir(log_dir)):
        m = LOG_TXT_PAT.match(fname)
        if not m: continue
        dmy, _ = m.groups()
        try:
            dt = parse_date(dmy)
        except:
            continue
        if dt < cutoff_txt:
            try:
                os.remove(os.path.join(log_dir, fname))
                print(f"[Temizlik] TXT silindi: {fname}")
            except Exception as e:
                print(f"[Uyarı] TXT silinemedi ({fname}): {e}")

    # ZIP tutma (son 2 hafta)
    keep_weeks = set()
    y0, w0 = iso_year_week(now_next_minute.date())
    prev = now_next_minute.date() - datetime.timedelta(weeks=1)
    y1, w1 = iso_year_week(prev)
    keep_weeks.update([(y0, w0), (y1, w1)])

    zip_pat = re.compile(r"^week_(\d{4})_(\d{1,2})_.+?(?:_part\d+)?\.zip$", re.IGNORECASE)
    for fname in list(os.listdir(log_dir)):
        if not (fname.endswith(".zip") and fname.startswith("week_")):
            continue
        m = zip_pat.match(fname)
        if not m: continue
        yy, ww = int(m.group(1)), int(m.group(2))
        if (yy, ww) not in keep_weeks:
            try:
                os.remove(os.path.join(log_dir, fname))
                print(f"[Temizlik] ZIP silindi: {fname}")
            except Exception as e:
                print(f"[Uyarı] ZIP silinemedi ({fname}): {e}")

# === BACKLOG (Geçmiş haftaları hemen gönder) ===
def iter_week_txts(lab: str, year: int, week: int) -> list[str]:
    """Belirli LAB ve ISO (yıl, hafta) için TXT dosyalarının isimlerini döndürür."""
    out = []
    for fname in os.listdir(log_dir):
        m = LOG_TXT_PAT.match(fname)
        if not m: 
            continue
        dmy, lab_name = m.groups()
        if lab_name != lab:
            continue
        dt = parse_date(dmy)
        y, w = iso_year_week(dt)
        if (y, w) == (year, week):
            out.append(fname)
    return sorted(out, key=lambda n: parse_date(LOG_TXT_PAT.match(n).group(1)))

def ensure_excels_for_week(lab: str, year: int, week: int):
    """Belirli hafta/LAB için mevcut TXT’lerden eksik kalan Excel’leri üretir."""
    txt_names = iter_week_txts(lab, year, week)
    for fname in txt_names:
        txt_path = os.path.join(log_dir, fname)
        ensure_excel_for_txt(txt_path, lab)

def delete_week_excels(lab: str, year: int, week: int):
    """Belirli hafta/LAB için üretilmiş Excel’leri siler."""
    for fname in list(os.listdir(log_dir)):
        if not (fname.endswith(".xlsx") and fname.startswith("sensor_log_")):
            continue
        m = LOG_TXT_PAT.match(fname.replace(".xlsx", ".txt"))
        if not m: 
            continue
        dmy, lab_name = m.groups()
        if lab_name != lab:
            continue
        dt = parse_date(dmy)
        y, w = iso_year_week(dt)
        if (y, w) == (year, week):
            try:
                os.remove(os.path.join(log_dir, fname))
            except Exception as e:
                print(f"[Uyarı] Excel silinemedi ({fname}): {e}")

def send_backlog_completed_weeks():
    """
    ŞU ANKİ hafta HARİÇ, klasörde görünen tüm hafta/LAB kombinasyonları için:
    - TXT→XLSX tamamla
    - XLSX’leri ziple
    - ZIP’leri mail at
    - Başarılı gönderimden sonra o haftanın Excel’lerini sil
    """
    today = datetime.date.today()
    cur_y, cur_w = iso_year_week(today)

    # Klasörü tarayıp görünen tüm (lab, y, w) setini çıkar
    seen = set()
    for fname in os.listdir(log_dir):
        m = LOG_TXT_PAT.match(fname)
        if not m:
            continue
        dmy, lab = m.groups()
        try:
            dt = parse_date(dmy)
        except:
            continue
        y, w = iso_year_week(dt)
        if (y, w) == (cur_y, cur_w):
            continue  # içinde bulunduğumuz hafta değil
        seen.add((lab, y, w))

    if not seen:
        print("[Backlog] Geçmişe dönük gönderilecek hafta yok.")
        return

    any_zip = False
    for (lab, y, w) in sorted(seen, key=lambda t: (t[1], t[2], t[0])):
        ensure_excels_for_week(lab, y, w)
        zips = zip_week_excels(lab, y, w)
        if zips:
            any_zip = True
            print(f"[Backlog] {lab} için {y}-W{w} zip hazır: {len(zips)} parça.")

    if any_zip:
        send_pending_zip_files()
        for (lab, y, w) in sorted(seen, key=lambda t: (t[1], t[2], t[0])):
            delete_week_excels(lab, y, w)

# === ANA DÖNGÜ ===
try:
    while True:
        # 1) Sensörleri oku ve günlük TXT’ye yaz
        for lab_name, config in sensors.items():
            values = read_values(config['ip'], config['registers'])
            log_data(lab_name, values)

        # 2) Backlog'u bir kereye mahsus çalıştır: geçmiş haftaları HEMEN gönder
        global BACKLOG_PROCESSED
        if not BACKLOG_PROCESSED:
            send_backlog_completed_weeks()
            BACKLOG_PROCESSED = True

        # 3) Zamanlama işlerini işlet
        now = datetime.datetime.now()
        next_minute = (now + datetime.timedelta(minutes=1)).replace(second=0, microsecond=0)

        # Gün sonu: 00:00'da bir önceki günün TXT→Excel
        if next_minute.hour == 0 and next_minute.minute == 0:
            yesterday = now.date()
            daily_finalize_to_excel(yesterday)

        # Hafta sonu: Pazar 00:00'da (dünün ISO haftası) zip + mail + excel cleanup
        weekly_finalize_and_send_if_sunday00(next_minute)

        # Arşiv temizlik: her gece 00:05
        if next_minute.hour == 0 and next_minute.minute == 5:
            retention_cleanup(next_minute)

        # Dakikanın başına kadar uyku
        time.sleep((next_minute - now).total_seconds())

except KeyboardInterrupt:
    print("Loglama durduruldu.")
